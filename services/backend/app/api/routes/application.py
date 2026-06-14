import io
import re
from datetime import datetime, timezone
from urllib.parse import urlparse, urlunparse

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from database_pkg.models import (
    Application,
    ApplicationStatus,
    ApplicationStatusHistory,
)
from database_pkg.models import (
    GeneratedLetter as DBGeneratedLetter,
)
from database_pkg.models import (
    JobDescription as DBJobDescription,
)
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from sqlmodel import func, select

from app.api.deps import CurrentUser, SessionDep
from app.api.validation.schemas import (
    CreateApplicationRequest,
    SaveApplicationRequest,
    SaveApplicationResponse,
    UpdateApplicationRequest,
    UpdateApplicationStatusRequest,
)

# Matches characters that cannot be serialized to XML (control chars in range 0x00-0x1F, except tab, LF, CR)
ILLEGAL_CHARACTERS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

router = APIRouter(tags=["application"])


@router.post("/save-application", response_model=SaveApplicationResponse)
async def save_application(
    request: SaveApplicationRequest, session: SessionDep, current_user: CurrentUser
):
    try:
        assert current_user.id is not None
        parsed = urlparse(request.job_url)
        netloc = parsed.netloc.lower()
        path = parsed.path.rstrip("/") or "/"
        normalized_netloc = netloc.removeprefix("www.")
        normalized_url = urlunparse(
            (
                parsed.scheme,
                normalized_netloc,
                path,
                parsed.params,
                parsed.query,
                parsed.fragment,
            )
        ).rstrip("/")
        www_url = urlunparse(
            (
                parsed.scheme,
                f"www.{normalized_netloc}",
                path,
                parsed.params,
                parsed.query,
                parsed.fragment,
            )
        ).rstrip("/")
        url_variants = set([request.job_url, normalized_url, www_url])
        url_variants.add(normalized_url + "/")
        url_variants.add(www_url + "/")
        job_id = None
        m = re.search(r"linkedin\.com/jobs/view/(\d+)", request.job_url)
        if not m:
            m = re.search(r"currentJobId=(\d+)", request.job_url)
        if m:
            job_id = m.group(1)
            url_variants.add(f"https://www.linkedin.com/jobs/view/{job_id}/")
            url_variants.add(f"https://linkedin.com/jobs/view/{job_id}/")
            url_variants.add(
                f"https://www.linkedin.com/jobs/collections/recommended/?currentJobId={job_id}"
            )
            url_variants.add(
                f"https://linkedin.com/jobs/collections/recommended/?currentJobId={job_id}"
            )
        statement = select(DBJobDescription).where(
            DBJobDescription.url.in_(list(url_variants))  # type: ignore[operator]
        )
        existing_job = session.exec(statement).first()
        if existing_job:
            job_description = existing_job
            assert job_description.id is not None
        else:
            job_description = DBJobDescription(
                url=request.job_url,
                full_description=request.job_description,
                requirements=request.job_requirements,
                job_title=request.job_title,
                company=request.job_company,
                source=request.job_source,
            )
            session.add(job_description)
            session.commit()
            session.refresh(job_description)
            assert job_description.id is not None
        generated_letters_data = [
            {
                "model": letter.model,
                "letter": letter.letter,
                "timestamp": letter.timestamp,
            }
            for letter in request.generated_letters
        ]
        generated_letter = DBGeneratedLetter(
            user_id=current_user.id, generated_letters=generated_letters_data
        )
        session.add(generated_letter)
        session.commit()
        session.refresh(generated_letter)
        assert generated_letter.id is not None
        selected_letter = request.generated_letters[request.selected_letter_index]
        cover_letter_final = {
            "model": selected_letter.model,
            "timestamp": selected_letter.timestamp,
            "body": request.cover_letter_body,
        }
        existing_app_statement = select(Application).where(
            Application.user_id == current_user.id,
            Application.job_description_id == job_description.id,
        )
        existing_application = session.exec(existing_app_statement).first()
        if existing_application:
            existing_application.generated_letter_id = generated_letter.id
            existing_application.header = request.header
            existing_application.cover_letter_final = cover_letter_final
            existing_application.updated_at = datetime.now(timezone.utc)
            session.add(existing_application)
            session.commit()
            session.refresh(existing_application)
            assert existing_application.id is not None
            application = existing_application
        else:
            application = Application(
                user_id=current_user.id,
                job_description_id=job_description.id,
                generated_letter_id=generated_letter.id,
                header=request.header,
                cover_letter_final=cover_letter_final,
                status=ApplicationStatus.APPLIED,
            )
            session.add(application)
            session.commit()
            session.refresh(application)
            assert application.id is not None
            history = ApplicationStatusHistory(
                application_id=application.id,
                old_status=None,
                new_status=ApplicationStatus.APPLIED,
                notes="Initial application creation",
            )
            session.add(history)
            session.commit()
        assert application.id is not None
        assert job_description.id is not None
        assert generated_letter.id is not None
        return SaveApplicationResponse(
            success=True,
            application_id=application.id,
            job_description_id=job_description.id,
            generated_letter_id=generated_letter.id,
        )
    except Exception as e:
        session.rollback()
        raise HTTPException(
            status_code=500, detail=f"Failed to save application: {str(e)}"
        )


@router.post("/application")
async def create_manual_application(
    request: CreateApplicationRequest, session: SessionDep, current_user: CurrentUser
):
    try:
        assert current_user.id is not None
        job_description = None
        if request.job_url:
            statement = select(DBJobDescription).where(
                DBJobDescription.url == request.job_url
            )
            job_description = session.exec(statement).first()
        if not job_description:
            job_description = DBJobDescription(
                url=request.job_url
                or f"manual-{datetime.now(timezone.utc).timestamp()}",
                job_title=request.job_title,
                company=request.company,
                full_description=f"Manually created application for {request.job_title} at {request.company}",
                requirements=[],
                source="Manual Entry",
            )
            session.add(job_description)
            session.commit()
            session.refresh(job_description)
            assert job_description.id is not None
        assert job_description.id is not None
        application = Application(
            user_id=current_user.id,
            job_description_id=job_description.id,
            status=ApplicationStatus(request.status),
            notes=request.notes,
            cover_letter_final={
                "model": "Manual",
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "body": request.cover_letter_body,
            },
            header={},
        )
        session.add(application)
        session.commit()
        session.refresh(application)
        assert application.id is not None
        history = ApplicationStatusHistory(
            application_id=application.id,
            old_status=None,
            new_status=application.status,
            notes="Manual application creation",
        )
        session.add(history)
        session.commit()
        return {"success": True, "application_id": application.id}
    except Exception as e:
        session.rollback()
        raise HTTPException(
            status_code=500, detail=f"Failed to create manual application: {str(e)}"
        )


@router.get("/applications")
async def get_user_applications(
    session: SessionDep,
    current_user: CurrentUser,
    limit: int = 50,
    skip: int = 0,
    include_details: bool = False,
    sort_by: str = "created_at",
    sort_order: str = "desc",
    company: str | None = None,
):
    try:
        count_stmt = (
            select(func.count(Application.id))  # type: ignore[arg-type]
            .join(DBJobDescription)
            .where(Application.user_id == current_user.id)
        )
        if company:
            count_stmt = count_stmt.where(
                DBJobDescription.company.ilike(f"%{company}%")  # type: ignore[union-attr]
            )
        total = session.exec(count_stmt).one()
        sort_map = {
            "created_at": Application.created_at,
            "company": DBJobDescription.company,
            "status": Application.status,
            "job_title": DBJobDescription.job_title,
        }
        sort_col = sort_map.get(sort_by, Application.created_at)
        from sqlmodel import asc, desc

        order_expr = desc(sort_col) if sort_order == "desc" else asc(sort_col)
        statement = (
            select(Application, DBJobDescription)
            .join(DBJobDescription)
            .where(Application.user_id == current_user.id)
        )
        if company:
            statement = statement.where(
                DBJobDescription.company.ilike(f"%{company}%")  # type: ignore[union-attr]
            )
        statement = statement.order_by(order_expr).offset(skip).limit(limit)
        results = session.exec(statement).all()
        applications = []
        for app, job_desc in results:
            app_dict = {
                "id": app.id,
                "job_title": job_desc.job_title,
                "company": job_desc.company,
                "job_url": job_desc.url,
                "status": app.status.value,
                "notes": app.notes,
                "created_at": app.created_at.isoformat(),
            }
            if include_details:
                app_dict["header"] = app.header
                app_dict["cover_letter_final"] = app.cover_letter_final
                app_dict["job_description"] = job_desc.full_description
                app_dict["requirements"] = job_desc.requirements
            applications.append(app_dict)
        return {"applications": applications, "total": total}
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to fetch applications: {str(e)}"
        )


@router.get("/applications/companies", response_model=list[str])
async def get_user_companies(session: SessionDep, current_user: CurrentUser):
    try:
        statement = (
            select(DBJobDescription.company)
            .join(Application)
            .where(Application.user_id == current_user.id)
            .distinct()
            .order_by(DBJobDescription.company)
        )
        companies = session.exec(statement).all()
        return [c for c in companies if c]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/applications/export")
async def export_applications_excel(
    session: SessionDep,
    current_user: CurrentUser,
    format: str = "xlsx",
):
    """Export all user applications to an Excel (.xlsx) or CSV file."""
    try:
        statement = (
            select(Application, DBJobDescription)
            .join(DBJobDescription)
            .where(Application.user_id == current_user.id)
            .order_by(Application.created_at.desc())  # type: ignore[union-attr]
        )
        results = session.exec(statement).all()

        rows = []
        for app, job_desc in results:
            cover_letter_body = ""
            if app.cover_letter_final and "body" in app.cover_letter_final:
                cover_letter_body = app.cover_letter_final["body"] or ""
            requirements = ""
            if job_desc.requirements:
                requirements = ", ".join(job_desc.requirements)
            rows.append(
                {
                    "date": app.created_at.strftime("%Y-%m-%d"),
                    "company": job_desc.company or "",
                    "job_title": job_desc.job_title or "",
                    "status": app.status.value if app.status else "",
                    "job_url": job_desc.url or "",
                    "notes": app.notes or "",
                    "requirements": requirements,
                    "cover_letter": cover_letter_body,
                    "job_description": job_desc.full_description or "",
                }
            )

        # ── CSV branch ─────────────────────────────────────────────────────
        if format.lower() == "csv":
            import csv

            output = io.StringIO()
            writer = csv.DictWriter(
                output,
                fieldnames=[
                    "date",
                    "company",
                    "job_title",
                    "status",
                    "job_url",
                    "notes",
                    "requirements",
                    "cover_letter",
                    "job_description",
                ],
            )
            writer.writeheader()
            writer.writerows(rows)
            csv_bytes = output.getvalue().encode("utf-8-sig")
            today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
            filename = f"vite-a-job-applications-{today}.csv"
            return StreamingResponse(
                io.BytesIO(csv_bytes),
                media_type="text/csv",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                    "Content-Length": str(len(csv_bytes)),
                },
            )

        # ── XLSX branch ────────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "My Applications"  # type: ignore[union-attr]

        headers = [
            "Date",
            "Company",
            "Job Title",
            "Status",
            "Job URL",
            "Notes",
            "Requirements",
            "Cover Letter",
            "Job Description",
        ]
        keys = [
            "date",
            "company",
            "job_title",
            "status",
            "job_url",
            "notes",
            "requirements",
            "cover_letter",
            "job_description",
        ]

        # Header row styling — sky-blue background, white bold text
        header_fill = PatternFill(
            start_color="0369A1", end_color="0369A1", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=False
        )

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)  # type: ignore[union-attr]
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Freeze the header row
        ws.freeze_panes = "A2"  # type: ignore[union-attr]

        # Data rows
        row_alignment = Alignment(vertical="top", wrap_text=True)
        for row_idx, row_data in enumerate(rows, start=2):
            # Alternate row background for readability
            if row_idx % 2 == 0:
                row_fill = PatternFill(
                    start_color="EFF6FF", end_color="EFF6FF", fill_type="solid"
                )
            else:
                row_fill = PatternFill(fill_type=None)

            for col_idx, key in enumerate(keys, start=1):
                val = row_data[key]
                if isinstance(val, str):
                    val = ILLEGAL_CHARACTERS_RE.sub(
                        lambda m: f"\\x{ord(m.group(0)):02x}", val
                    )
                cell = ws.cell(row=row_idx, column=col_idx, value=val)  # type: ignore[union-attr]
                cell.alignment = row_alignment
                cell.fill = row_fill
                cell.font = Font(name="Calibri", size=10)

        # Set column widths
        col_widths = {
            1: 14,  # Date
            2: 22,  # Company
            3: 30,  # Job Title
            4: 12,  # Status
            5: 45,  # Job URL
            6: 30,  # Notes
            7: 40,  # Requirements
            8: 60,  # Cover Letter
            9: 60,  # Job Description
        }
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width  # type: ignore[union-attr]

        # Set row height for data rows (tall for wrapped text)
        for row_idx in range(2, len(rows) + 2):
            ws.row_dimensions[row_idx].height = 60  # type: ignore[union-attr]

        # Stream the workbook
        output_bytes = io.BytesIO()
        wb.save(output_bytes)
        output_bytes.seek(0)
        xlsx_content = output_bytes.read()

        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        filename = f"vite-a-job-applications-{today}.xlsx"
        return StreamingResponse(
            io.BytesIO(xlsx_content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Content-Length": str(len(xlsx_content)),
            },
        )
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to export applications: {str(e)}"
        )


@router.get("/application/{application_id}/details")
async def get_application_details(
    application_id: int, session: SessionDep, current_user: CurrentUser
):
    try:
        statement = (
            select(Application, DBJobDescription)
            .join(
                DBJobDescription,
                Application.job_description_id == DBJobDescription.id,  # type: ignore[arg-type]
            )
            .where(
                Application.id == application_id, Application.user_id == current_user.id
            )
        )
        result = session.exec(statement).first()
        if not result:
            raise HTTPException(status_code=404, detail="Application not found")
        app, job_desc = result
        return {
            "id": app.id,
            "job_title": job_desc.job_title,
            "company": job_desc.company,
            "status": app.status.value,
            "job_url": job_desc.url,
            "header": app.header,
            "cover_letter_final": app.cover_letter_final,
            "notes": app.notes,
            "job_description": job_desc.full_description,
            "requirements": job_desc.requirements,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to fetch application details: {str(e)}"
        )


@router.get("/application/check-duplicate")
async def check_duplicate_application(
    job_url: str, session: SessionDep, current_user: CurrentUser
):
    try:
        parsed = urlparse(job_url)
        netloc = parsed.netloc.lower()
        path = parsed.path.rstrip("/") or "/"
        normalized_netloc = netloc.removeprefix("www.")
        normalized_url = urlunparse(
            (
                parsed.scheme,
                normalized_netloc,
                path,
                parsed.params,
                parsed.query,
                parsed.fragment,
            )
        ).rstrip("/")
        www_url = urlunparse(
            (
                parsed.scheme,
                f"www.{normalized_netloc}",
                path,
                parsed.params,
                parsed.query,
                parsed.fragment,
            )
        ).rstrip("/")
        url_variants = set([job_url, normalized_url, www_url])
        url_variants.add(normalized_url + "/")
        url_variants.add(www_url + "/")
        job_id = None
        m = re.search(r"linkedin\.com/jobs/view/(\d+)", job_url)
        if not m:
            m = re.search(r"currentJobId=(\d+)", job_url)
        if m:
            job_id = m.group(1)
            url_variants.add(f"https://www.linkedin.com/jobs/view/{job_id}/")
            url_variants.add(f"https://linkedin.com/jobs/view/{job_id}/")
            url_variants.add(
                f"https://www.linkedin.com/jobs/collections/recommended/?currentJobId={job_id}"
            )
            url_variants.add(
                f"https://linkedin.com/jobs/collections/recommended/?currentJobId={job_id}"
            )
        url_variants = list(url_variants)
        statement = (
            select(Application, DBJobDescription)
            .join(
                DBJobDescription,
                Application.job_description_id == DBJobDescription.id,  # type: ignore[arg-type]
            )
            .where(
                Application.user_id == current_user.id,
                DBJobDescription.url.in_(url_variants),  # type: ignore[operator]
            )
        )
        result = session.exec(statement).first()
        if not result:
            return {"is_duplicate": False, "existing_application": None}
        app, job_desc = result
        cover_letter_body = None
        if app.cover_letter_final and "body" in app.cover_letter_final:
            cover_letter_body = app.cover_letter_final["body"]
        return {
            "is_duplicate": True,
            "existing_application": {
                "id": app.id,
                "job_title": job_desc.job_title,
                "company": job_desc.company,
                "status": app.status.value,
                "notes": app.notes,
                "cover_letter_body": cover_letter_body,
                "created_at": app.created_at.isoformat(),
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to check duplicate application: {str(e)}"
        )


@router.patch("/application/{application_id}/status")
async def update_application_status(
    application_id: int,
    request: UpdateApplicationStatusRequest,
    session: SessionDep,
    current_user: CurrentUser,
):
    try:
        statement = select(Application).where(
            Application.id == application_id, Application.user_id == current_user.id
        )
        application = session.exec(statement).first()
        if not application:
            raise HTTPException(status_code=404, detail="Application not found")
        try:
            new_status = ApplicationStatus(request.status)
        except ValueError:
            raise HTTPException(
                status_code=400, detail=f"Invalid status: {request.status}"
            )
        application.status = new_status
        session.add(application)
        session.commit()
        session.refresh(application)
        return {"success": True, "status": application.status.value}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(
            status_code=500, detail=f"Failed to update status: {str(e)}"
        )


@router.delete("/application/{application_id}")
async def delete_application(
    application_id: int, session: SessionDep, current_user: CurrentUser
):
    try:
        statement = select(Application).where(
            Application.id == application_id, Application.user_id == current_user.id
        )
        application = session.exec(statement).first()
        if not application:
            raise HTTPException(status_code=404, detail="Application not found")
        letter_id = application.generated_letter_id
        history_statement = select(ApplicationStatusHistory).where(
            ApplicationStatusHistory.application_id == application_id
        )
        history_records = session.exec(history_statement).all()
        for record in history_records:
            session.delete(record)
        related_letters_stmt = select(DBGeneratedLetter).where(
            DBGeneratedLetter.application_id == application_id
        )
        related_letters = session.exec(related_letters_stmt).all()
        session.delete(application)
        session.flush()
        deleted_letter_ids = set()
        for letter in related_letters:
            deleted_letter_ids.add(letter.id)
            session.delete(letter)
        if letter_id and letter_id not in deleted_letter_ids:
            refs_stmt = select(func.count(Application.id)).where(  # type: ignore[arg-type]
                Application.generated_letter_id == letter_id,
                Application.id != application_id,
            )
            remaining_refs = session.exec(refs_stmt).one()
            if remaining_refs == 0:
                specific_letter = session.get(DBGeneratedLetter, letter_id)
                if specific_letter and specific_letter.user_id == current_user.id:
                    session.delete(specific_letter)
        session.commit()
        return {"success": True, "message": "Application deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(
            status_code=500, detail=f"Failed to delete application: {str(e)}"
        )


@router.patch("/application/{application_id}")
async def update_application(
    application_id: int,
    request: UpdateApplicationRequest,
    session: SessionDep,
    current_user: CurrentUser,
):
    try:
        statement = (
            select(Application, DBJobDescription)
            .join(
                DBJobDescription,
                Application.job_description_id == DBJobDescription.id,  # type: ignore[arg-type]
            )
            .where(
                Application.id == application_id, Application.user_id == current_user.id
            )
        )
        result = session.exec(statement).first()
        if not result:
            raise HTTPException(status_code=404, detail="Application not found")
        application, job_description = result
        assert application.id is not None
        if request.status is not None:
            try:
                new_status = ApplicationStatus(request.status)
                if new_status != application.status:
                    history = ApplicationStatusHistory(
                        application_id=application.id,
                        old_status=application.status,
                        new_status=new_status,
                        notes=request.notes
                        if request.notes
                        else "Status manual update",
                    )
                    session.add(history)
                    application.status = new_status
            except ValueError:
                raise HTTPException(
                    status_code=400, detail=f"Invalid status: {request.status}"
                )
        if request.notes is not None:
            application.notes = request.notes
        if request.header is not None:
            application.header = request.header
        if request.cover_letter_body is not None:
            new_final = dict(application.cover_letter_final)
            new_final["body"] = request.cover_letter_body
            application.cover_letter_final = new_final
        application.updated_at = datetime.now(timezone.utc)
        job_updated = False
        if request.job_title is not None:
            job_description.job_title = request.job_title
            job_updated = True
        if request.company is not None:
            job_description.company = request.company
            job_updated = True
        if job_updated:
            job_description.updated_at = datetime.now(timezone.utc)
            session.add(job_description)
        session.add(application)
        session.commit()
        session.refresh(application)
        return {"success": True, "application_id": application.id}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(
            status_code=500, detail=f"Failed to update application: {str(e)}"
        )


@router.get("/application/{application_id}/history")
async def get_application_history(
    application_id: int, session: SessionDep, current_user: CurrentUser
):
    try:
        statement = select(Application).where(
            Application.id == application_id, Application.user_id == current_user.id
        )
        application = session.exec(statement).first()
        if not application:
            raise HTTPException(status_code=404, detail="Application not found")
        history_statement = (
            select(ApplicationStatusHistory)
            .where(ApplicationStatusHistory.application_id == application_id)
            .order_by(ApplicationStatusHistory.created_at.desc())  # type: ignore[union-attr]
        )
        history = session.exec(history_statement).all()
        return history
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to fetch history: {str(e)}"
        )
